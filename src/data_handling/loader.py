"""
Handles loading and preprocessing of input data from CSV or Excel files.

This module is responsible for:
- Reading data using pandas, with a "smart read" capability to handle files
  with an unknown number of trailing empty rows.
- Standardizing column names based on configurable input profiles.
- Initializing new columns required by the pipeline (e.g., for status tracking,
  extracted data, and run identifiers).
- Performing initial normalization of phone numbers if a designated phone
  number column is present in the input.
"""
import csv # For smart CSV reading
import logging
import uuid # For RunID
from typing import Optional, List, Dict, Any, Union, Iterable

import pandas as pd
from openpyxl import load_workbook # For smart Excel reading

# Import AppConfig directly. Its __init__ handles .env loading.
# If this import fails, it's a critical setup error for the application.
from ..core.config import AppConfig

# Configure logging.
# The setup_logging() function might rely on environment variables that are
# loaded when AppConfig is instantiated.
try:
    from ..core.logging_config import setup_logging
    # AppConfig() is instantiated globally in config.py if needed by other modules,
    # or when an instance is created. Here, we just ensure logging is set up.
    setup_logging()
    logger = logging.getLogger(__name__)
except ImportError:
    # Fallback basic logging configuration if core.logging_config is unavailable.
    # This might happen during isolated testing or if there's a setup issue.
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    logger = logging.getLogger(__name__)
    logger.warning( # Changed from info to warning as this is a fallback
        "Basic logging configured for loader.py due to missing "
        "core.logging_config or its dependencies. This is a fallback."
    )


def _is_row_empty(row_values: Iterable[Any]) -> bool:
    """
    Checks if all values in a given row are effectively empty.

    An empty value is defined as None, an empty string, or a string
    containing only whitespace.

    Args:
        row_values: An iterable (e.g., list, tuple) of values in a row.

    Returns:
        True if all values in the row are empty, False otherwise.
        Returns True if row_values itself is None or an empty iterable.
    """
    if not row_values: # Handles case where row_values might be None or an empty list/tuple
        return True
    return all(pd.isna(value) or (isinstance(value, str) and not value.strip()) for value in row_values)


def load_and_preprocess_data(
    file_path: str,
    app_config_instance: Optional[AppConfig] = None
) -> Optional[pd.DataFrame]:
    """
    Loads data from a CSV or Excel file, standardizes column names, initializes
    new pipeline columns, and applies initial phone number normalization.

    This function supports "smart reading" for files where the exact number of
    data rows is unknown. It stops reading after encountering a configurable
    number of consecutive empty rows.

    Args:
        file_path (str): The path to the input CSV or Excel file.
        app_config_instance (Optional[AppConfig]): An optional instance of
            AppConfig. If not provided, a new one will be instantiated.
            This allows for passing a pre-configured AppConfig instance,
            useful for testing or specific runtime configurations.

    Returns:
        pd.DataFrame | None: The processed DataFrame with standardized and
        new columns. Returns None if a critical error occurs during
        loading (e.g., file not found, unsupported file type).
        Returns an empty DataFrame if the input file is empty or
        contains no valid data after applying skip/read limits.
    """
    current_config_instance: AppConfig
    if app_config_instance:
        current_config_instance = app_config_instance
    else:
        # Instantiate AppConfig if not provided; it handles .env loading.
        current_config_instance = AppConfig()

    skip_rows_val: Optional[int] = None
    nrows_val: Optional[int] = None
    # Default for smart read, will be overridden by config if available.
    consecutive_empty_rows_to_stop: int = 3

    # Load row range configurations if they exist in AppConfig
    # These attributes might not exist if AppConfig is a minimal version or old.
    if hasattr(current_config_instance, 'skip_rows_config'):
        skip_rows_val = current_config_instance.skip_rows_config
    if hasattr(current_config_instance, 'nrows_config'):
        nrows_val = current_config_instance.nrows_config
    if hasattr(current_config_instance, 'consecutive_empty_rows_to_stop'):
        consecutive_empty_rows_to_stop = current_config_instance.consecutive_empty_rows_to_stop

    log_message_parts = []
    if skip_rows_val is not None:
        # Clarify that skip_rows_val is 0-indexed from the start of the file (header is row 0)
        log_message_parts.append(f"skipping first {skip_rows_val} rows (0-indexed from file start, header is row 0)")
    if nrows_val is not None:
        log_message_parts.append(f"reading a maximum of {nrows_val} data rows (after any skipped rows)")

    # Smart read is active if nrows_val is not set (open-ended read) AND
    # consecutive_empty_rows_to_stop is a positive number.
    smart_read_active = (nrows_val is None and consecutive_empty_rows_to_stop > 0)
    if smart_read_active:
        log_message_parts.append(f"smart read active (will stop after {consecutive_empty_rows_to_stop} consecutive empty data rows)")

    if log_message_parts:
        logger.info(f"Data loading configuration: {', '.join(log_message_parts)}.")
    else:
        logger.info("No specific row range configured; attempting to load all rows (or smart read if enabled by default).")

    # pandas_skiprows_arg is for when not using smart read, or for the initial skip
    # before smart reading begins. It refers to lines in the file (1-indexed for skiprows list).
    # Pandas `skiprows` parameter:
    # - int: number of lines to skip from start of file (0 means no lines, header is line 0)
    # - list-like: 0-indexed line numbers to skip. So [0] skips header. [1] skips first data row.
    # Our `skip_rows_val` is intended to mean "number of rows to skip *after* the header".
    # So, if skip_rows_val = 1, we want pandas to skip file line 1 (the first data row).
    pandas_skiprows_arg: Union[int, List[int]]
    if skip_rows_val is not None and skip_rows_val > 0:
        # To skip `skip_rows_val` data rows, we need to skip file lines 1 to `skip_rows_val`.
        pandas_skiprows_arg = list(range(1, skip_rows_val + 1))
    else:
        pandas_skiprows_arg = 0 # Skip no lines after the header (header=0 means header is row 0)

    df: Optional[pd.DataFrame] = None

    try:
        logger.info(f"Attempting to load data from: {file_path}")

        if smart_read_active:
            logger.info(f"Smart read enabled. Max consecutive empty rows to stop: {consecutive_empty_rows_to_stop}")
            header: Optional[List[str]] = None
            data_rows: List[List[Any]] = []
            empty_row_counter = 0

            # actual_data_rows_to_skip refers to 0-indexed data rows (rows *after* the header)
            actual_data_rows_to_skip = skip_rows_val if skip_rows_val is not None else 0

            if file_path.endswith(('.xls', '.xlsx')):
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = workbook.active

                if sheet is None:
                    logger.warning(f"Excel file {file_path} does not have an active sheet or is empty. Returning empty DataFrame.")
                    header = None # Ensure header is None for empty DataFrame creation
                    data_rows = [] # Ensure data_rows is empty
                else:
                    rows_iter = sheet.iter_rows()

                    # 1. Read header from the first row of the sheet
                    try:
                        header_row_values = next(rows_iter)
                        header = [str(cell.value) if cell.value is not None else '' for cell in header_row_values]
                        logger.info(f"Excel header read: {header}")
                    except StopIteration: # Handles empty sheet
                        logger.warning(f"Excel file {file_path} seems empty (no header row found).")
                        # header will remain None, data_rows empty.

                    # Only proceed to read data rows if header was successfully read
                    if header is not None:
                        # 2. Skip initial data rows (after header)
                        for _ in range(actual_data_rows_to_skip):
                            try:
                                next(rows_iter)
                            except StopIteration:
                                logger.info(f"Reached end of Excel file while skipping initial {actual_data_rows_to_skip} data rows.")
                                break

                        # 3. Read data rows with empty row detection
                        for row_idx, row_values_tuple in enumerate(rows_iter):
                            current_row_values = [cell.value for cell in row_values_tuple]
                            if _is_row_empty(current_row_values):
                                empty_row_counter += 1
                                if empty_row_counter >= consecutive_empty_rows_to_stop:
                                    logger.info(f"Stopping Excel read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                    break
                            else:
                                empty_row_counter = 0 # Reset counter on non-empty row
                                data_rows.append(current_row_values)

                # Create DataFrame from collected header and data_rows
                if header: # If header was read
                    df = pd.DataFrame(data_rows, columns=header)
                    logger.info(f"Smart read from Excel resulted in {len(data_rows)} data rows.")
                elif not data_rows and header is None: # Handles case where sheet was None or truly empty
                    df = pd.DataFrame() # Create an empty DataFrame
                    logger.info("Smart read from Excel: sheet was None or empty, created empty DataFrame.")
                # else: # This case (header is None but data_rows has content) should ideally not occur.
                # df = pd.DataFrame(data_rows) # Fallback if header is somehow None but data exists

            elif file_path.endswith('.csv'):
                with open(file_path, mode='r', encoding='utf-8', newline='') as csvfile:
                    reader = csv.DictReader(csvfile)

                    # 1. Read header
                    try:
                        header = list(reader.fieldnames) if reader.fieldnames else None
                        logger.info(f"CSV header read: {header}")
                    except StopIteration: # Handles empty CSV
                        logger.warning(f"CSV file {file_path} seems empty (no header row).")
                        return pd.DataFrame() # Return empty DataFrame

                    # 2. Skip initial data rows
                    for _ in range(actual_data_rows_to_skip):
                        try:
                            next(reader)
                        except StopIteration:
                            logger.info(f"Reached end of CSV file while skipping initial {actual_data_rows_to_skip} data rows.")
                            break

                    # 3. Read data with empty row detection
                    for row_idx, current_row_values in enumerate(reader):
                        # csv.reader can yield empty lists for completely blank lines
                        is_empty = not current_row_values or _is_row_empty(current_row_values.values())

                        if is_empty:
                            empty_row_counter += 1
                            if empty_row_counter >= consecutive_empty_rows_to_stop:
                                logger.info(f"Stopping CSV read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                break
                        else:
                            empty_row_counter = 0
                            data_rows.append(list(current_row_values.values()))

                if header: # If header was read
                    df = pd.DataFrame(data_rows, columns=header)
                # else: # Should not be reached if header read was successful and file wasn't empty
                # df = pd.DataFrame(data_rows)
                logger.info(f"Smart read from CSV resulted in {len(data_rows)} data rows.")
            else:
                logger.error(f"Unsupported file type for smart read: {file_path}. Please use CSV or Excel.")
                return None
        else: # Standard pandas read (fixed range or smart read disabled)
            logger.info(f"Using standard pandas read. Pandas skiprows argument: {pandas_skiprows_arg}, nrows: {nrows_val}")
            # keep_default_na=False and na_filter=False to prevent pandas from interpreting empty strings as NaN
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val, keep_default_na=False, na_filter=False)
            elif file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val, keep_default_na=False, na_filter=False)
            else:
                logger.error(f"Unsupported file type: {file_path}. Please use CSV or Excel.")
                return None

        if df is None: # Should only occur if smart read was attempted for an unsupported file type
            logger.error(f"DataFrame is None after loading attempt for {file_path}. This indicates an issue with the loading logic or unsupported file for smart read.")
            return None

        logger.info(f"Columns loaded: {df.columns.tolist() if df is not None and not df.empty else 'N/A (DataFrame is None or empty)'}")

        if df.empty:
            logger.warning(f"Loaded DataFrame from {file_path} is empty. This could be due to an empty input file, all rows being skipped, or smart read stopping early.")
            # If df is empty, we still want to ensure essential columns are present for later stages.
            # The new_columns loop later will add them if they don't exist.

        # --- Post-loading processing: Apply input profile for column renaming and add new pipeline columns ---
        active_profile_name = current_config_instance.input_file_profile_name
        profile_mappings = current_config_instance.INPUT_COLUMN_PROFILES.get(active_profile_name)

        if not profile_mappings:
            logger.error(f"Input profile '{active_profile_name}' not found in AppConfig.INPUT_COLUMN_PROFILES. Falling back to 'default' profile.")
            active_profile_name = "default" # Attempt to use a default profile
            profile_mappings = current_config_instance.INPUT_COLUMN_PROFILES.get("default")
            if not profile_mappings: # This should ideally not happen if "default" is always defined in AppConfig
                 logger.error("Critical: Default input profile ('default') not found in AppConfig. Cannot map columns.")
                 return pd.DataFrame() # Return empty DataFrame as a fallback

        # Create rename map only for columns present in the DataFrame
        actual_rename_map = {k: v for k, v in profile_mappings.items() if not k.startswith('_') and k in df.columns}

        if actual_rename_map:
             df.rename(columns=actual_rename_map, inplace=True)
        logger.info(f"DataFrame columns after renaming (using profile: '{active_profile_name}'): {df.columns.tolist()}")

        # Define and initialize new columns required by the pipeline
        new_columns = [
            "NormalizedGivenPhoneNumber", "ScrapingStatus",
            "Overall_VerificationStatus", "Original_Number_Status",
            "Primary_Number_1", "Primary_Type_1", "Primary_SourceURL_1",
            "Secondary_Number_1", "Secondary_Type_1", "Secondary_SourceURL_1",
            "Secondary_Number_2", "Secondary_Type_2", "Secondary_SourceURL_2",
            "RunID", "TargetCountryCodes"
        ]

        current_run_id = str(uuid.uuid4()) # Generate a unique RunID for this processing batch

        for col in new_columns:
            if col not in df.columns:
                if col == "RunID":
                    df[col] = current_run_id
                elif col == "TargetCountryCodes":
                    # Initialize with default target countries; robust for empty df
                    df[col] = pd.Series([["DE", "AT", "CH"] for _ in range(len(df))] if not df.empty else [], dtype=object)
                elif col in ["ScrapingStatus", "Overall_VerificationStatus", "Original_Number_Status"]:
                    df[col] = "Pending" # Default status
                elif col.startswith("Primary_") or col.startswith("Secondary_"):
                    df[col] = None # Initialize phone/type/source columns as None
                else:
                    df[col] = None # Default for other new columns
        
        logger.info(f"Successfully loaded and structured data from {file_path}. DataFrame shape: {df.shape}")

        return df
    except FileNotFoundError:
        logger.error(f"Error: The file {file_path} was not found.")
        return None
    except pd.errors.EmptyDataError: # This might be caught by smart read logic earlier for empty files
        logger.error(f"Error: The file {file_path} is empty (pandas EmptyDataError).")
        # Return an empty DataFrame and None for the phone column name, consistent with other error paths.
        return pd.DataFrame()
    except Exception as e:
        logger.error(f"An unexpected error occurred while loading data from {file_path}: {e}", exc_info=True)
        return None