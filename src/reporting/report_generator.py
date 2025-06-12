"""
Generates various Excel-based reports for the contact pipeline.

This module contains functions to create specific reports such as:
- Row Attrition Report: Details input rows that did not yield contacts and why.
- Canonical Domain Summary Report: Summarizes processing outcomes for each unique canonical domain.
- Augmented Input Report: Enhances the original input file with processed data (e.g., verified numbers).
- Detailed LLM Extractions Report: Provides a granular view of all phone numbers extracted by the LLM.
- Summary Report: A high-level overview of phone validation statuses.
- Tertiary (Contact Focused) Report: Focuses on top contact information per company.

Each report generation function handles data formatting, Excel file creation,
and column width adjustments for readability.
"""
import os
import logging
import pandas as pd
from openpyxl.utils import get_column_letter
# from datetime import datetime # Not directly used in this module after cleanup
import json
# from pathlib import Path # Not directly used in this module after cleanup
from typing import List, Dict, Optional, Any, Tuple, Counter as TypingCounter
from collections import Counter

# from src.core.config import AppConfig # AppConfig is not directly used.
from src.utils.helpers import get_input_canonical_url

# It's good practice to get the logger for the current module
logger_module = logging.getLogger(__name__)


def write_row_attrition_report(
    run_id: str,
    attrition_data: List[Dict[str, Any]],
    output_dir: str,
    canonical_domain_journey_data: Dict[str, Dict[str, Any]],
    input_to_canonical_map: Dict[str, Optional[str]],
    logger: logging.Logger  # Keep passed logger for consistency if other modules do this
) -> int:
    """
    Writes the collected row attrition data to an Excel file.

    This report details input rows that failed to yield contact information,
    along with reasons and links to canonical domain processing outcomes.
    Columns are auto-adjusted for width.

    Args:
        run_id (str): The unique identifier for the current pipeline run.
        attrition_data (List[Dict[str, Any]]): A list of dictionaries, where each
            dictionary represents a row that did not yield a contact.
        output_dir (str): The directory where the Excel file will be saved.
        canonical_domain_journey_data (Dict[str, Dict[str, Any]]): Data mapping
            canonical domains to their processing journey details. Used for linking.
        input_to_canonical_map (Dict[str, Optional[str]]): Mapping of input URLs
            to their determined canonical URLs. Used for linking.
        logger (logging.Logger): Logger instance for logging messages.

    Returns:
        int: The number of rows written to the report, or 0 if an error occurred
             or no data was provided.
    """
    if not attrition_data:
        logger.info("No data for row attrition report. Skipping file creation.")
        return 0

    report_filename = f"row_attrition_report_{run_id}.xlsx"
    report_path = os.path.join(output_dir, report_filename)
    
    # Create a deep copy for modification to avoid unintended changes to the original list of dicts
    attrition_data_copy = [dict(row) for row in attrition_data]

    # Enhance attrition data with derived fields for the report
    for row_data in attrition_data_copy:
        given_url = row_data.get("GivenURL")
        row_data["Derived_Input_CanonicalURL"] = get_input_canonical_url(given_url) # Helper to get canonical form of input URL
        
        final_processed_domain = row_data.get("Relevant_Canonical_URLs") # This might be a list or single string
        # Ensure Final_Processed_Canonical_Domain is a single representative domain or None
        if isinstance(final_processed_domain, list) and final_processed_domain:
            row_data["Final_Processed_Canonical_Domain"] = final_processed_domain[0] # Take the first if it's a list
        elif isinstance(final_processed_domain, str) and pd.notna(final_processed_domain) and final_processed_domain != "N/A":
            row_data["Final_Processed_Canonical_Domain"] = final_processed_domain
        else:
            row_data["Final_Processed_Canonical_Domain"] = None

        # Attempt to link to the canonical domain journey data
        link_to_outcome = None
        current_final_domain_for_link = row_data["Final_Processed_Canonical_Domain"]
        if current_final_domain_for_link: # Check if a valid domain string exists
            if current_final_domain_for_link in canonical_domain_journey_data:
                link_to_outcome = current_final_domain_for_link
            else:
                # Fallback: try to find via input_to_canonical_map if direct link fails
                input_url_key = str(given_url) if given_url is not None else "None_GivenURL_Input"
                mapped_canonical = input_to_canonical_map.get(input_url_key)
                if mapped_canonical and mapped_canonical in canonical_domain_journey_data:
                    link_to_outcome = mapped_canonical
                    logger.debug(
                        f"AttritionReport: Linked via input_to_canonical_map. GivenURL: {given_url} "
                        f"-> MappedCanonical: {mapped_canonical}"
                    )
                else:
                    logger.debug(
                        f"AttritionReport: Could not find domain '{current_final_domain_for_link}' "
                        f"(from GivenURL: {given_url}, MappedCanonical: {mapped_canonical}) "
                        f"in canonical_domain_journey_data for linking."
                    )
        row_data["Link_To_Canonical_Domain_Outcome"] = link_to_outcome

    report_df = pd.DataFrame(attrition_data_copy)

    # Define the desired column order for the report
    columns_order = [
        "InputRowID", "CompanyName", "GivenURL",
        "Derived_Input_CanonicalURL",
        "Final_Processed_Canonical_Domain", # The single representative domain
        "Link_To_Canonical_Domain_Outcome",
        "Final_Row_Outcome_Reason", "Determined_Fault_Category",
        "Relevant_Canonical_URLs", # Original field, might be list
        "LLM_Error_Detail_Summary",
        "Input_CompanyName_Total_Count",
        "Input_CanonicalURL_Total_Count",
        "Is_Input_CompanyName_Duplicate",
        "Is_Input_CanonicalURL_Duplicate",
        "Is_Input_Row_Considered_Duplicate",
        "Timestamp_Of_Determination"
    ]
    
    # Ensure all specified columns exist in the DataFrame, adding them with None if missing
    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None
    report_df = report_df[columns_order] # Reorder/select columns

    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Attrition_Report')
            worksheet = writer.sheets['Attrition_Report']
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(report_df.columns):
                series_data = report_df.iloc[:, col_idx] # Use iloc for positional access
                max_val_len = 0
                if not series_data.empty:
                    # Ensure all data is string before calculating length
                    lengths = series_data.astype(str).map(len)
                    if not lengths.empty:
                        max_val_len = lengths.max()
                
                column_header_len = len(str(col_name))
                # Add a small buffer for padding
                adjusted_width = max(max_val_len, column_header_len) + 2
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"Row attrition report successfully saved to {report_path}")
        return len(report_df)
    except Exception as e:
        logger.error(f"Failed to write row attrition report to {report_path}: {e}", exc_info=True)
        return 0


def write_canonical_domain_summary_report(
    run_id: str,
    domain_journey_data: Dict[str, Dict[str, Any]],
    output_dir: str,
    logger: logging.Logger # Keep passed logger
) -> int:
    """
    Writes the canonical domain journey data to an Excel file.

    This report summarizes the processing steps and outcomes for each unique
    canonical domain encountered during the pipeline run.

    Args:
        run_id (str): The unique identifier for the current pipeline run.
        domain_journey_data (Dict[str, Dict[str, Any]]): A dictionary where keys
            are canonical domains and values are dictionaries of their processing journey.
        output_dir (str): The directory where the Excel file will be saved.
        logger (logging.Logger): Logger instance for logging messages.

    Returns:
        int: The number of rows written to the report, or 0 if an error occurred
             or no data was provided.
    """
    if not domain_journey_data:
        logger.info("No data for canonical domain summary report. Skipping file creation.")
        return 0

    report_filename = f"canonical_domain_processing_summary_{run_id}.xlsx"
    report_path = os.path.join(output_dir, report_filename)

    # Convert the dictionary of dictionaries to a list of dictionaries for DataFrame creation
    report_data_list = []
    for domain, data_dict in domain_journey_data.items():
        row = {"Canonical_Domain": domain}
        row.update(data_dict) # Add all other data from the inner dictionary
        report_data_list.append(row)
    
    if not report_data_list: # Should be redundant due to the initial check, but good for safety
        logger.info("Formatted report_data_list is empty. Skipping canonical domain summary report.")
        return 0
        
    report_df = pd.DataFrame(report_data_list)

    # Define the desired column order
    columns_order = [
        "Canonical_Domain", "Input_Row_IDs", "Input_CompanyNames", "Input_GivenURLs",
        "Pathful_URLs_Attempted_List", "Overall_Scraper_Status_For_Domain",
        "Total_Pages_Scraped_For_Domain", "Scraped_Pages_Details_Aggregated", # This is likely a Counter
        "Regex_Candidates_Found_For_Any_Pathful", "LLM_Calls_Made_For_Domain",
        "LLM_Total_Raw_Numbers_Extracted", "LLM_Total_Consolidated_Numbers_Found",
        "LLM_Consolidated_Number_Types_Summary", # This is likely a Counter
        "LLM_Processing_Error_Encountered_For_Domain",
        "LLM_Error_Messages_Aggregated", "Final_Domain_Outcome_Reason",
        "Primary_Fault_Category_For_Domain"
    ]

    # Ensure all specified columns exist, adding them with None if missing
    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None
            logger.debug(f"Column '{col}' was not found in canonical_domain_summary_report DataFrame and was initialized to None.") # Changed to debug
    report_df = report_df[columns_order] # Reorder/select columns

    # Convert list/set fields to comma-separated strings for better Excel readability
    list_like_columns_to_join = [
        "Input_Row_IDs", "Input_CompanyNames", "Input_GivenURLs",
        "Pathful_URLs_Attempted_List", "LLM_Error_Messages_Aggregated"
    ]
    for col_name in list_like_columns_to_join:
        if col_name in report_df.columns:
            report_df[col_name] = report_df[col_name].apply(
                lambda x: ", ".join(sorted(list(map(str, x)))) if isinstance(x, (set, list)) else x
            )
    
    # Convert Counter objects to JSON strings for Excel readability
    counter_columns_to_json = ["Scraped_Pages_Details_Aggregated", "LLM_Consolidated_Number_Types_Summary"]
    for col_name in counter_columns_to_json:
        if col_name in report_df.columns:
            report_df[col_name] = report_df[col_name].apply(
                lambda x: json.dumps(dict(x)) if isinstance(x, (Counter, TypingCounter)) else x
            )

    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Canonical_Domain_Summary')
            worksheet = writer.sheets['Canonical_Domain_Summary']
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(report_df.columns):
                series_data = report_df[col_name] # Direct access by name
                max_val_len = 0
                if not series_data.empty:
                    lengths = series_data.astype(str).map(len)
                    if not lengths.empty:
                         max_val_len = lengths.max()
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 5 # Increased buffer slightly
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"Canonical domain summary report successfully saved to {report_path}")
        return len(report_df)
    except Exception as e:
        logger.error(f"Failed to write canonical domain summary report to {report_path}: {e}", exc_info=True)
        return 0

